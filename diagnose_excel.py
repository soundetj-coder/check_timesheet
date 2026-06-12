#!/usr/bin/env python3
"""
Excel残業申請シート診断スクリプト

実行方法:
    python diagnose_excel.py

config.yaml の overtime 設定を使ってExcelファイルを開き、
実際に読み取れる値を表示します。
"違反なし" になる原因の特定に使ってください。
"""

import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
import yaml

CONFIG_PATH = Path(__file__).parent / "config.yaml"


def main():
    with open(CONFIG_PATH, encoding="utf-8") as f:
        config = yaml.safe_load(f)

    ot_cfg = config.get("overtime", {})
    excel_path = ot_cfg.get("excel_path", "").strip()

    if not excel_path:
        print("overtime.excel_path が設定されていません")
        sys.exit(1)

    if not Path(excel_path).exists():
        print(f"Excelファイルが見つかりません:\n  {excel_path}")
        sys.exit(1)

    print(f"Excelファイル: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print(f"シート一覧: {wb.sheetnames}")

    # config から対象シートを特定
    today = date.today()
    from datetime import timedelta
    import jpholiday
    d = today - timedelta(days=1)
    while d.weekday() >= 5 or jpholiday.is_holiday(d):
        d -= timedelta(days=1)
    prev_biz_day = d
    print(f"前営業日: {prev_biz_day}")

    sheet_fmt = ot_cfg.get("sheet_name_format", "{month}月")
    sheet_name = sheet_fmt.format(month=prev_biz_day.month, year=prev_biz_day.year)
    print(f"対象シート名（config設定）: '{sheet_name}'")

    if sheet_name not in wb.sheetnames:
        print(f"\n★ シート '{sheet_name}' が見つかりません！")
        print(f"  利用可能なシート: {wb.sheetnames}")
        print("  → config.yaml の sheet_name_format を修正してください")
        sys.exit(1)

    ws = wb[sheet_name]
    print(f"シートサイズ: {ws.max_column}列 × {ws.max_row}行\n")

    day_row = ot_cfg.get("header_day_row", 13)
    sub_row = ot_cfg.get("subheader_row", 15)
    name_col_letter = ot_cfg.get("name_column", "C")
    name_col = column_index_from_string(name_col_letter)
    data_start = ot_cfg.get("data_start_row", 16)
    target_day = prev_biz_day.day

    print(f"=" * 60)
    print(f"設定値: header_day_row={day_row}, subheader_row={sub_row}")
    print(f"       name_column={name_col_letter}, data_start_row={data_start}")
    print(f"       対象日: {prev_biz_day.month}月{target_day}日")
    print(f"=" * 60)

    # 行 day_row の内容
    print(f"\n【行{day_row}の内容（日付行として設定）】")
    non_empty = 0
    for col in range(1, min(ws.max_column + 1, 100)):
        v = ws.cell(row=day_row, column=col).value
        if v is not None:
            non_empty += 1
            col_letter = get_column_letter(col)
            print(f"  {col_letter}{day_row}: {repr(v)}  (型: {type(v).__name__})")
    if non_empty == 0:
        print(f"  ★ 行{day_row} は全セルが空です！")
        print(f"  → header_day_row の値が違う可能性があります")
        print(f"  → 前後の行（{day_row-2}〜{day_row+2}）を確認してください")
        print()
        for r in range(max(1, day_row - 2), day_row + 3):
            row_vals = []
            for col in range(1, min(ws.max_column + 1, 20)):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    row_vals.append(f"{get_column_letter(col)}: {repr(v)}")
            if row_vals:
                print(f"  行{r}: {', '.join(row_vals[:10])}")

    # 行 sub_row の内容
    print(f"\n【行{sub_row}の内容（申請/打刻/時間行として設定）】")
    non_empty = 0
    for col in range(1, min(ws.max_column + 1, 100)):
        v = ws.cell(row=sub_row, column=col).value
        if v is not None:
            non_empty += 1
            col_letter = get_column_letter(col)
            print(f"  {col_letter}{sub_row}: {repr(v)}")
    if non_empty == 0:
        print(f"  ★ 行{sub_row} は全セルが空です！")

    # 対象日の列を探す
    print(f"\n【{target_day}日の申請列を探す】")
    current_day = None
    found_app_col = -1
    for col in range(1, ws.max_column + 1):
        dv = ws.cell(row=day_row, column=col).value
        # _parse_day_number と同じロジック
        parsed = None
        if isinstance(dv, bool):
            pass
        elif isinstance(dv, int) and 1 <= dv <= 31:
            parsed = dv
        elif isinstance(dv, float) and 1 <= int(dv) <= 31:
            parsed = int(dv)
        elif isinstance(dv, datetime):
            parsed = dv.day
        elif isinstance(dv, date):
            parsed = dv.day
        elif isinstance(dv, str):
            import re
            m = re.search(r"\d{1,2}", dv)
            if m:
                i = int(m.group(0))
                if 1 <= i <= 31:
                    parsed = i
        if parsed is not None:
            current_day = parsed

        sub = ws.cell(row=sub_row, column=col).value
        if current_day == target_day and isinstance(sub, str) and "申請" in sub:
            found_app_col = col
            print(f"  → 申請列発見: {get_column_letter(col)}{sub_row} (列番号={col})")
            break

    if found_app_col < 0:
        print(f"  ★ {target_day}日の申請列が見つかりませんでした")

    # 氏名列の内容
    print(f"\n【列{name_col_letter}の氏名一覧（行{data_start}以降）】")
    names_found = []
    for row in range(data_start, min(ws.max_row + 1, data_start + 100)):
        v = ws.cell(row=row, column=name_col).value
        if v is not None and str(v).strip():
            names_found.append((row, str(v)))
            print(f"  行{row}: {repr(str(v))}")
    if not names_found:
        print(f"  ★ 列{name_col_letter}（行{data_start}以降）に氏名が見つかりません")

    # 申請列の内容（見つかった場合）
    if found_app_col > 0 and names_found:
        print(f"\n【{target_day}日の申請時刻（列{get_column_letter(found_app_col)}）】")
        for row, name in names_found:
            v = ws.cell(row=row, column=found_app_col).value
            print(f"  行{row} {name}: {repr(v)}  (型: {type(v).__name__ if v is not None else 'None'})")

    print("\n診断完了")


if __name__ == "__main__":
    main()
