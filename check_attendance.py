#!/usr/bin/env python3
"""
タイムカード自動確認・通知スクリプト
Windowsタスクスケジューラで毎朝10:00に実行する
"""

import json
import re
import smtplib
import socket
import sys
from datetime import date, datetime, time, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from urllib.error import URLError
from urllib.request import Request, urlopen

import jpholiday
import yaml
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

CONFIG_PATH = Path(__file__).parent / "config.yaml"
LOG_DIR = Path(__file__).parent / "logs"

# 氏名列と認識するヘッダー文字列
_NAME_HEADERS = {"氏名", "名前", "社員名", "氏　名", "氏　　名", "社員氏名", "name"}


def load_config() -> dict:
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return yaml.safe_load(f)


def log(msg: str) -> None:
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {msg}", flush=True)


def _hhmm_to_minutes(s: str) -> int:
    """"HH:MM" を分に換算する。不正な値は -1。"""
    try:
        h, m = s.split(":")
        return int(h) * 60 + int(m)
    except (ValueError, AttributeError):
        return -1


# ---------------------------------------------------------------------------
# 平日判定・前営業日計算
# ---------------------------------------------------------------------------

def is_business_day(d: date) -> bool:
    """平日（土日・日曜・祝日以外）の場合 True を返す。"""
    if d.weekday() >= 5:  # 5=土曜、6=日曜
        return False
    if jpholiday.is_holiday(d):
        return False
    return True


def skip_reason(d: date) -> str:
    """スキップ理由の文字列を返す（ログ出力用）。"""
    if d.weekday() == 5:
        return "土曜日"
    if d.weekday() == 6:
        return "日曜日"
    holiday_name = jpholiday.is_holiday(d)
    return f"祝日（{holiday_name}）"


def get_previous_business_day(today: date) -> date:
    """今日の1日前から順に遡り、最初に見つかった平日を返す。"""
    d = today - timedelta(days=1)
    while not is_business_day(d):
        d -= timedelta(days=1)
    return d


# ---------------------------------------------------------------------------
# Playwright ヘルパー
# ---------------------------------------------------------------------------

def _click_list_button(page) -> None:
    """「一覧表示」ボタンをクリックする。見つからない場合は RuntimeError を送出。"""
    try:
        page.click('input[value="一覧表示"]', timeout=10_000)
        return
    except Exception:
        pass
    try:
        page.get_by_text("一覧表示").click(timeout=10_000)
    except Exception as exc:
        raise RuntimeError(f"一覧表示ボタンが見つかりませんでした: {exc}") from exc


def _switch_month(page, year: int, month: int) -> None:
    """月プルダウンを指定年月に切り替え、一覧表示を再クリックしてテーブルを待つ。"""
    month_value = f"{year:04d}-{month:02d}"
    try:
        page.select_option('select[name="month"]', month_value)
    except Exception as exc:
        raise RuntimeError(
            f"月プルダウン (select[name='month']) の操作に失敗しました。\n"
            f"  実際のHTMLのname属性を確認して調整してください。\n"
            f"  (値: {month_value}, エラー: {exc})"
        ) from exc

    _click_list_button(page)

    try:
        page.wait_for_selector("table", timeout=15_000)
    except PlaywrightTimeoutError as exc:
        raise RuntimeError(
            f"{year}年{month}月のテーブル表示待機でタイムアウトしました。"
        ) from exc


def _get_main_table_rows(page) -> list:
    """現在のページから最も行数の多いテーブルの行リストを返す。"""
    tables = page.query_selector_all("table")
    if not tables:
        return []
    main_table = max(tables, key=lambda t: len(t.query_selector_all("tr")))
    return main_table.query_selector_all("tr")


# ---------------------------------------------------------------------------
# スクレイピング
# ---------------------------------------------------------------------------

def scrape_attendance(
    config: dict, prev_biz_day: date
) -> tuple[list[dict], list[dict]]:
    """今日の打刻データと前営業日の全社員の打刻データを1セッションで取得する。"""
    url = config["url"]
    today = date.today()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(ignore_https_errors=True)
        page = context.new_page()

        try:
            page.goto(url, wait_until="networkidle", timeout=30_000)
        except PlaywrightTimeoutError:
            log("ページ読み込みタイムアウト。取得済みコンテンツで続行します。")

        try:
            _click_list_button(page)
        except RuntimeError as exc:
            log(str(exc))
            browser.close()
            return [], []

        try:
            page.wait_for_selector("table", timeout=15_000)
        except PlaywrightTimeoutError:
            log("テーブルの表示を待機中にタイムアウトしました。")
            browser.close()
            return [], []

        today_rows = _get_main_table_rows(page)
        if not today_rows:
            log("テーブルが見つかりませんでした。")
            browser.close()
            return [], []

        today_dt = datetime.combine(today, datetime.min.time())
        today_employees = _parse_rows(
            today_rows, _find_day_column(today_rows, today_dt), config
        )

        prev_in_different_month = (
            prev_biz_day.year != today.year or prev_biz_day.month != today.month
        )

        if prev_in_different_month:
            log(
                f"前営業日 ({prev_biz_day}) は前月のため、"
                f"{prev_biz_day.year}年{prev_biz_day.month}月に切り替えます。"
            )
            try:
                _switch_month(page, prev_biz_day.year, prev_biz_day.month)
                prev_rows = _get_main_table_rows(page)
            except RuntimeError as exc:
                log(f"月切り替えに失敗しました: {exc}")
                prev_rows = []
        else:
            prev_rows = today_rows

        prev_day_attendance = (
            _get_day_attendance(prev_rows, prev_biz_day) if prev_rows else []
        )

        browser.close()

    return today_employees, prev_day_attendance


def _get_day_attendance(rows: list, target_day: date) -> list[dict]:
    """指定日の列から全社員の出退勤時刻を抽出する。"""
    target_dt = datetime.combine(target_day, datetime.min.time())
    col_index = _find_day_column(rows, target_dt)

    if col_index < 0:
        log(f"{target_day.month}/{target_day.day} の列が見つかりませんでした。")
        return []

    name_col = _find_name_column(rows)
    result = []

    for row in rows[1:]:
        cells = row.query_selector_all("td, th")
        if not cells:
            continue

        name_idx = name_col if name_col < len(cells) else 0
        name = cells[name_idx].inner_text().strip()
        if not name:
            continue

        clock_in = clock_out = ""
        if col_index < len(cells):
            cell_text = cells[col_index].inner_text().strip()
            times = re.findall(r"\d{1,2}:\d{2}", cell_text)
            clock_in = times[0] if times else ""
            clock_out = times[1] if len(times) >= 2 else ""

        result.append(
            {
                "name": name,
                "clock_in": clock_in,
                "clock_out": clock_out,
                "date": target_day,
            }
        )

    return result


def _find_day_column(rows: list, today: datetime) -> int:
    """ヘッダー行から当日の列インデックスを返す。見つからない場合は -1。"""
    if not rows:
        return -1

    target = str(today.day)
    patterns = [
        target,
        f"{today.month}/{target}",
        f"{today.month:02d}/{int(target):02d}",
        f"{target}日",
    ]

    headers = rows[0].query_selector_all("th, td")
    for i, header in enumerate(headers):
        text = header.inner_text().strip()
        if text in patterns:
            return i
        if re.search(rf"(?<![\d]){re.escape(target)}(?![\d])", text):
            return i

    return -1


def _find_name_column(rows: list) -> int:
    """ヘッダー行から氏名列のインデックスを返す。見つからない場合は 0（先頭列）。"""
    if not rows:
        return 0

    headers = rows[0].query_selector_all("th, td")
    for i, header in enumerate(headers):
        text = header.inner_text().strip()
        if text in _NAME_HEADERS or any(h in text for h in _NAME_HEADERS):
            return i

    log("氏名列が特定できず先頭列を使用します。")
    return 0


def _parse_rows(rows: list, day_col_index: int, config: dict) -> list[dict]:
    """データ行を走査して社員リストを生成する。"""
    employees = []
    name_col_index = _find_name_column(rows)

    for row in rows[1:]:
        cells = row.query_selector_all("td, th")
        if not cells:
            continue

        name_idx = name_col_index if name_col_index < len(cells) else 0
        name = cells[name_idx].inner_text().strip()
        if not name:
            continue

        if 0 < day_col_index < len(cells):
            cell = cells[day_col_index]
            status, clock_in, clock_out = _analyze_cell(
                cell.inner_text().strip(),
                cell.inner_html(),
                config,
            )
        else:
            status, clock_in, clock_out = "未打刻", "", ""

        employees.append(
            {"name": name, "status": status, "clock_in": clock_in, "clock_out": clock_out}
        )

    return employees


# ---------------------------------------------------------------------------
# 打刻状況の判定
# ---------------------------------------------------------------------------

def _analyze_cell(cell_text: str, cell_html: str, config: dict) -> tuple[str, str, str]:
    """セルテキストとHTMLから (status, clock_in, clock_out) を返す。"""
    if not cell_text:
        return "未打刻", "", ""

    times = re.findall(r"\d{1,2}:\d{2}", cell_text)
    if not times:
        return "未打刻", "", ""

    clock_in = times[0]
    clock_out = times[1] if len(times) >= 2 else ""

    is_overtime = bool(clock_out) and "+" in cell_text
    is_late = _is_late(clock_in, cell_html, config)

    if not clock_out:
        status = "遅刻・退勤未打刻" if is_late else "出勤のみ（退勤未打刻）"
    elif is_late and is_overtime:
        status = "遅刻・残業"
    elif is_late:
        status = "遅刻"
    elif is_overtime:
        status = "残業"
    else:
        status = "正常"

    return status, clock_in, clock_out


def _is_late(clock_in: str, cell_html: str, config: dict) -> bool:
    """出勤時刻が遅刻かどうか判定する。"""
    html_lower = cell_html.lower()
    if (
        "color:red" in html_lower
        or "color: red" in html_lower
        or 'class="red"' in html_lower
        or "color=#ff" in html_lower
        or 'style="color:red' in html_lower
    ):
        return True

    late_threshold = config.get("late_threshold", "09:30")
    try:
        in_time = datetime.strptime(clock_in, "%H:%M").time()
        threshold = datetime.strptime(late_threshold, "%H:%M").time()
        return in_time > threshold
    except ValueError:
        return False


# ---------------------------------------------------------------------------
# 残業事前申請チェック（Excel参照）
# ---------------------------------------------------------------------------

def _excel_value_to_hhmm(value) -> str:
    """Excelセルの値を "HH:MM" 文字列に変換する。空・不明は空文字。"""
    if value is None:
        return ""
    if isinstance(value, str):
        m = re.search(r"\d{1,2}:\d{2}", value)
        return m.group(0) if m else ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, timedelta):
        total = int(value.total_seconds())
        return f"{total // 3600:02d}:{(total % 3600) // 60:02d}"
    if isinstance(value, (int, float)):
        # Excelの時刻シリアル値（1日=1.0 の小数）
        total_minutes = round(float(value) * 24 * 60)
        return f"{(total_minutes // 60) % 24:02d}:{total_minutes % 60:02d}"
    return ""


def _parse_day_number(v):
    """セル値から日番号（1〜31）を抽出する。見つからない場合は None。"""
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v if 1 <= v <= 31 else None
    if isinstance(v, float):
        i = int(v)
        return i if 1 <= i <= 31 else None
    # datetime は date のサブクラスなので先にチェックする
    if isinstance(v, datetime):
        return v.day
    if isinstance(v, date):
        return v.day
    if isinstance(v, str):
        m = re.search(r"\d{1,2}", v)
        if m:
            i = int(m.group(0))
            return i if 1 <= i <= 31 else None
    return None


def _find_application_column(ws, target_day: int, ot_cfg: dict) -> int:
    """指定日の「申請」列番号（1始まり）を返す。見つからない場合は -1。

    13行（日番号）はマージセルのため、値は左端セルにのみ入る。
    カラムを走査しながら直近の日番号を保持し、15行が「申請」の列を探す。
    """
    from openpyxl.utils import get_column_letter

    day_row = ot_cfg.get("header_day_row", 13)
    sub_row = ot_cfg.get("subheader_row", 15)
    current_day = None
    found_days = {}  # デバッグ用: {day: col}

    for col in range(1, ws.max_column + 1):
        dv = ws.cell(row=day_row, column=col).value
        parsed = _parse_day_number(dv)
        if parsed is not None:
            current_day = parsed
            if parsed not in found_days:
                found_days[parsed] = col

        sub = ws.cell(row=sub_row, column=col).value
        if (
            current_day == target_day
            and isinstance(sub, str)
            and "申請" in sub
        ):
            log(f"  → {target_day}日の申請列: {get_column_letter(col)}{sub_row}")
            return col

    # 見つからなかった場合のデバッグ情報
    log(f"  行{day_row}で検出された日番号: {sorted(found_days.items())}")
    if target_day not in found_days:
        log(f"  ※ {target_day}日が行{day_row}に存在しません（行番号設定を確認してください）")
    else:
        log(f"  ※ {target_day}日は列{get_column_letter(found_days[target_day])}にありますが、行{sub_row}に「申請」が見つかりません")
    return -1


def _normalize_name(name: str) -> str:
    """氏名を比較用に正規化する。

    ・タイムカード側: 「山田　太郎（ヤマダタロウ）」のようにフリガナが付く場合がある
    ・Excel側:       「山田　太郎」（漢字のみ）
    カッコ（半角・全角）内のフリガナを除去し、スペース（半角・全角）も除去して統一する。
    """
    # （フリガナ）や (フリガナ) を除去
    name = re.sub(r'[（(][^）)]*[）)]', '', name)
    return name.replace(" ", "").replace("　", "").strip()


def _find_employee_row(ws, name: str, name_col: int, data_start_row: int) -> int:
    """氏名列を走査して一致する社員の行番号を返す。見つからない場合は -1。"""
    target = _normalize_name(name)
    for row in range(data_start_row, ws.max_row + 1):
        v = ws.cell(row=row, column=name_col).value
        if v and _normalize_name(str(v)) == target:
            return row
    return -1


def check_overtime_applications(
    prev_overtime: list[dict], prev_biz_day: date, config: dict
) -> list[dict]:
    """前営業日に残業した社員について、Excelの事前申請と照合する。"""
    ot_cfg = config.get("overtime", {})
    if not ot_cfg.get("enabled", False):
        return []
    if not prev_overtime:
        return []

    excel_path = ot_cfg.get("excel_path", "").strip()
    if not excel_path:
        log("overtime.excel_path が未設定のため残業申請チェックをスキップします。")
        return []

    try:
        import openpyxl
        from openpyxl.utils import column_index_from_string
    except ImportError:
        log("openpyxl がインストールされていません。`pip install openpyxl` を実行してください。")
        return []

    if not Path(excel_path).exists():
        log(f"Excelファイルが見つかりません: {excel_path}")
        return []

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as exc:
        log(f"Excelファイルの読み込みに失敗しました: {exc}")
        return []

    sheet_name = ot_cfg.get("sheet_name_format", "{month}月").format(
        month=prev_biz_day.month, year=prev_biz_day.year
    )
    if sheet_name not in wb.sheetnames:
        log(f"シート '{sheet_name}' が見つかりません。利用可能: {wb.sheetnames}")
        return []
    ws = wb[sheet_name]

    log(f"Excel照合: シート='{sheet_name}', 対象日={prev_biz_day.month}/{prev_biz_day.day}")
    app_col = _find_application_column(ws, prev_biz_day.day, ot_cfg)
    if app_col < 0:
        log(f"{prev_biz_day.month}/{prev_biz_day.day} の「申請」列がシート上に見つかりませんでした。")
        return []

    name_col = column_index_from_string(ot_cfg.get("name_column", "C"))
    data_start_row = ot_cfg.get("data_start_row", 16)

    # 集計単位（分）。申請19:30 かつ unit_minutes=15 なら 19:44 まで許容
    unit_minutes = ot_cfg.get("unit_minutes", 15)
    tolerance = unit_minutes - 1

    violations = []
    for emp in prev_overtime:
        name = emp["name"]
        clock_out = emp["clock_out"]

        row = _find_employee_row(ws, name, name_col, data_start_row)
        if row < 0:
            log(f"  シート上に氏名なし: '{name}' (正規化後: '{_normalize_name(name)}')")
            violations.append(
                {"name": name, "clock_out": clock_out, "applied": "", "reason": "シート上に氏名なし"}
            )
            continue

        applied = _excel_value_to_hhmm(ws.cell(row=row, column=app_col).value)
        log(f"  {name}: 退勤={clock_out}, 申請={applied if applied else '(なし)'}")
        if not applied:
            violations.append(
                {"name": name, "clock_out": clock_out, "applied": "", "reason": "申請なし"}
            )
        elif _hhmm_to_minutes(clock_out) > _hhmm_to_minutes(applied) + tolerance:
            # 退勤が「申請時刻 + (集計単位-1)分」を超えた場合のみ超過とみなす
            # 例: 申請19:30, unit_minutes=15 → 19:44まで許容、19:45から超過
            violations.append(
                {
                    "name": name,
                    "clock_out": clock_out,
                    "applied": applied,
                    "reason": "申請時間超過",
                }
            )

    return violations


# ---------------------------------------------------------------------------
# 通知メッセージ生成
# ---------------------------------------------------------------------------

def _apply_morning_mode(employees: list[dict]) -> list[dict]:
    """朝チェックモード時、退勤未打刻を含むステータスを再分類する。"""
    mapping = {
        "出勤のみ（退勤未打刻）": "正常",
        "遅刻・退勤未打刻": "遅刻",
    }
    return [{**e, "status": mapping.get(e["status"], e["status"])} for e in employees]


def format_message(
    employees: list[dict],
    prev_missing: list[dict],
    overtime_violations: list[dict],
    prev_biz_day: date,
    config: dict,
) -> str:
    """今日の打刻・前営業日の退勤未打刻・残業申請違反を整形した通知文を返す。"""
    today = datetime.now()
    date_str = today.strftime("%Y年%m月%d日")
    prev_date_str = f"{prev_biz_day.month}/{prev_biz_day.day}"

    morning_mode = config.get("morning_check_mode", True)
    if morning_mode:
        employees = _apply_morning_mode(employees)

    STATUS_LABELS = {
        "未打刻": "未打刻",
        "出勤のみ（退勤未打刻）": "退勤未打刻",
        "遅刻・退勤未打刻": "遅刻+退勤未打刻",
        "遅刻": "遅刻",
        "遅刻・残業": "遅刻+残業",
        "残業": "残業",
        "正常": "正常",
    }

    issues = [e for e in employees if e["status"] != "正常"]
    normal = [e for e in employees if e["status"] == "正常"]

    mode_label = "（朝チェックモード）" if morning_mode else ""
    lines = [
        f"【タイムカード確認レポート】{date_str}{mode_label}",
        f"確認時刻: {today.strftime('%H:%M')}  基準時刻: {config.get('late_threshold', '09:30')}",
        "=" * 44,
    ]

    if issues:
        lines.append(f"\n■ 要確認（今日） ({len(issues)}名)")
        for e in issues:
            label = STATUS_LABELS.get(e["status"], e["status"])
            times = f"  出勤:{e['clock_in']}" if e["clock_in"] else ""
            if e["clock_out"]:
                times += f"  退勤:{e['clock_out']}"
            lines.append(f"  [{label}] {e['name']}{times}")
    else:
        lines.append("\n■ 要確認（今日）: なし")

    if prev_missing:
        lines.append(
            f"\n■ 前営業日（{prev_date_str}）退勤打刻なし ({len(prev_missing)}名)"
        )
        for e in prev_missing:
            if e["clock_in"]:
                detail = f"出勤:{e['clock_in']} / 退勤打刻なし"
            else:
                detail = "出勤・退勤ともに打刻なし"
            lines.append(f"  ・{e['name']}：前営業日（{prev_date_str}）{detail}")
    else:
        lines.append(
            f"\n■ 前営業日（{prev_date_str}）退勤打刻: 全員確認済み"
        )

    if config.get("overtime", {}).get("enabled", False):
        if overtime_violations:
            lines.append(
                f"\n■ 前営業日（{prev_date_str}）残業申請チェック ({len(overtime_violations)}名)"
            )
            for v in overtime_violations:
                if v["reason"] == "申請時間超過":
                    detail = f"退勤{v['clock_out']} / 申請{v['applied']}（超過）"
                elif v["reason"] == "申請なし":
                    detail = f"退勤{v['clock_out']} / 残業申請なし"
                else:
                    detail = f"退勤{v['clock_out']} / {v['reason']}"
                lines.append(f"  ・{v['name']}：{detail}")
        else:
            lines.append(
                f"\n■ 前営業日（{prev_date_str}）残業申請チェック: 違反なし"
            )

    if normal:
        lines.append(f"\n■ 正常打刻（今日） ({len(normal)}名)")
        for e in normal:
            times = f"  出勤:{e['clock_in']}" if e["clock_in"] else ""
            if e["clock_out"]:
                times += f"  退勤:{e['clock_out']}"
            lines.append(f"  {e['name']}{times}")

    lines.append(f"\n合計: {len(employees)}名")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# メール送信
# ---------------------------------------------------------------------------

def send_email(message: str, config: dict) -> None:
    email_cfg = config["email"]
    today = datetime.now()
    subject = f"タイムカード確認レポート {today.strftime('%Y/%m/%d')}"

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = email_cfg["sender"]
    msg["To"] = ", ".join(email_cfg["recipients"])
    msg.attach(MIMEText(message, "plain", "utf-8"))

    smtp_server = email_cfg["smtp_server"]
    smtp_port = email_cfg["smtp_port"]
    sender = email_cfg["sender"]
    password = email_cfg.get("password", "")
    use_tls = email_cfg.get("use_tls", True)
    use_ssl = email_cfg.get("use_ssl", False)

    try:
        if use_ssl:
            with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
                if password:
                    server.login(sender, password)
                server.sendmail(sender, email_cfg["recipients"], msg.as_string())
        else:
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                if use_tls:
                    server.starttls()
                if password:
                    server.login(sender, password)
                server.sendmail(sender, email_cfg["recipients"], msg.as_string())
        log("メール送信完了")
    except socket.gaierror:
        log(
            f"メール送信エラー: SMTPサーバー '{smtp_server}' の名前解決に失敗しました。\n"
            f"  → config.yaml の smtp_server を正しいホスト名またはIPアドレスに変更してください。"
        )
        raise
    except ConnectionRefusedError:
        log(
            f"メール送信エラー: SMTPサーバー '{smtp_server}:{smtp_port}' に接続できませんでした。\n"
            f"  → smtp_server ・ smtp_port ・ use_tls/use_ssl の設定を確認してください。"
        )
        raise
    except smtplib.SMTPAuthenticationError:
        log(
            "メール送信エラー: SMTP認証に失敗しました。\n"
            "  → config.yaml の sender ・ password を確認してください。"
        )
        raise
    except Exception as exc:
        log(f"メール送信エラー: {exc}")
        raise


# ---------------------------------------------------------------------------
# Slack 通知
# ---------------------------------------------------------------------------

def send_slack(message: str, config: dict) -> None:
    webhook_url = config.get("slack_webhook_url", "").strip()
    if not webhook_url:
        log("slack_webhook_url が未設定のためスキップします。")
        return

    payload = json.dumps({"text": f"```\n{message}\n```"}).encode("utf-8")
    req = Request(
        webhook_url,
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        with urlopen(req, timeout=10) as resp:
            if resp.status == 200:
                log("Slack 通知完了")
            else:
                log(f"Slack 通知失敗: HTTP {resp.status}")
    except URLError as exc:
        log(f"Slack 送信エラー: {exc}")
        raise


# ---------------------------------------------------------------------------
# エントリポイント
# ---------------------------------------------------------------------------

def main() -> None:
    LOG_DIR.mkdir(exist_ok=True)
    log("タイムカード確認開始")

    today = date.today()
    if not is_business_day(today):
        log(f"本日は{skip_reason(today)}のため処理をスキップします。")
        sys.exit(0)

    prev_biz_day = get_previous_business_day(today)
    log(f"前営業日: {prev_biz_day.strftime('%Y-%m-%d (%A)')}")

    config = load_config()

    log("スクレイピング中...")
    today_employees, prev_attendance = scrape_attendance(config, prev_biz_day)

    if not today_employees:
        log("社員データを取得できませんでした。スクリプトを終了します。")
        sys.exit(1)

    log(f"{len(today_employees)} 名のデータを取得しました。")

    prev_missing = [e for e in prev_attendance if not e["clock_out"]]
    if prev_missing:
        log(f"前営業日退勤未打刻: {len(prev_missing)} 名")

    ot_cfg = config.get("overtime", {})
    threshold_min = _hhmm_to_minutes(ot_cfg.get("overtime_threshold", "18:30"))
    prev_overtime = [
        e
        for e in prev_attendance
        if e["clock_out"] and _hhmm_to_minutes(e["clock_out"]) >= threshold_min
    ]

    if prev_overtime:
        log(f"前営業日残業者: {len(prev_overtime)} 名 → Excelで事前申請を照合します。")
        for e in prev_overtime:
            log(f"  - {e['name']}: 退勤={e['clock_out']}")
    else:
        log("前営業日残業者: 0名")
        if prev_attendance:
            log(f"  前営業日打刻サンプル（最大5名、閾値={ot_cfg.get('overtime_threshold','18:30')}）:")
            for e in prev_attendance[:5]:
                log(f"  - {e['name']}: 出勤={e['clock_in']}, 退勤={e['clock_out']}")
        else:
            log("  前営業日の打刻データが0件です（タイムカード側でデータが取れていません）")

    overtime_violations = check_overtime_applications(prev_overtime, prev_biz_day, config)
    if overtime_violations:
        log(f"残業申請違反: {len(overtime_violations)} 名")

    message = format_message(
        today_employees, prev_missing, overtime_violations, prev_biz_day, config
    )
    log("\n" + message + "\n")

    failed: list[str] = []

    try:
        send_email(message, config)
    except Exception:
        failed.append("メール")

    try:
        send_slack(message, config)
    except Exception:
        failed.append("Slack")

    if failed:
        log(f"通知失敗: {', '.join(failed)}")
        sys.exit(1)

    log("タイムカード確認完了")


if __name__ == "__main__":
    main()
